from fastapi import UploadFile, HTTPException
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from models import User
import bcrypt  


# Función para hashear la contraseña
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt)
    return hashed.decode("utf-8")


# Función para importar Excel
async def import_excel(file: UploadFile, db: Session):
    workbook = load_workbook(file.file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))[1:]

    for row in rows:
        name, email, password, rol = row
        hashed_password = hash_password(password)
        
        # Crear usuario
        user = User(name=name, email=email, password=hashed_password, rol=rol)
        db.add(user)
    
    db.commit()
    return {"message": "Excel data imported successfully"}
